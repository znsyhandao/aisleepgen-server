# -*- coding: utf-8 -*-
"""Parse Mendeley XLSX for PSQI/diagnosis columns"""
import openpyxl

xlsx_path = r'D:\AISleepGen_Optimized\data\edf\mendeley_insomnia\PSG_Psycho_Normal.xlsx'
wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)')
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header_row = list(row)
            break
    print(f'  All columns headers: {header_row}')
    keywords = ['sleep', 'score', 'quality', 'pill', 'medication', 'diagnosis', 'psqi', 'isi']
    matches = [(j, h) for j, h in enumerate(header_row) if h and any(k in str(h).lower() for k in keywords)]
    print(f'  Health-related columns: {matches}')
    print()
wb.close()
