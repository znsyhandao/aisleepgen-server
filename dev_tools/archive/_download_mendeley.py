# -*- coding: utf-8 -*-
"""Download Mendeley insomnia dataset"""
import urllib.request, json, os, time

DATA_DIR = r'D:\AISleepGen_Optimized\data\edf\mendeley_insomnia'
os.makedirs(DATA_DIR, exist_ok=True)

url = 'https://data.mendeley.com/public-api/datasets/3hx58k232n'
req = urllib.request.Request(url, headers={
    'Accept': 'application/json',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0'
})
with urllib.request.urlopen(req, timeout=15) as resp:
    data = json.loads(resp.read())

print(f'Version: {data.get("version")}')
print(f'Name: {data.get("name")}')
print()

downloads = []
for f in data.get('files', []):
    cd = f.get('content_details', {})
    sz_mb = cd.get('size', f.get('size', 0)) / 1e6
    dl_url = cd.get('download_url', '')
    fname = f['filename']
    print(f'  [{fname}]  {sz_mb:.1f}MB')
    print(f'    download_url: {dl_url}')
    desc = f.get('description', '')[:120]
    if desc:
        print(f'    desc: {desc}')
    print()
    if dl_url:
        downloads.append((fname, dl_url, cd.get('size', 0)))

print(f'Total downloadable files: {len(downloads)}')
print()

# Download the XLSX first (small, demographics + PSQI)
for fname, dl_url, size in downloads:
    if not fname.endswith('.xlsx'):
        continue
    dst = os.path.join(DATA_DIR, fname)
    if os.path.exists(dst):
        print(f'  [SKIP] {fname} already exists')
        continue
    print(f'  Downloading {fname} ({size/1e6:.1f}MB)...')
    dreq = urllib.request.Request(dl_url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0'
    })
    with urllib.request.urlopen(dreq, timeout=120) as resp:
        with open(dst, 'wb') as f:
            f.write(resp.read())
    print(f'  [OK] {fname} downloaded')
    time.sleep(1)

# Check what's in the XLSX
import subprocess
xlsx_path = os.path.join(DATA_DIR, 'PSG_Psycho_Normal.xlsx')
if os.path.exists(xlsx_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f'\nSheet: {sheet_name}')
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                print(f'  Row {i}: {list(row)[:15]}')
                if i >= 3:
                    print(f'  ... ({ws.max_row} total rows)')
                    break
        wb.close()
    except ImportError:
        print('openpyxl not installed, try pip install openpyxl')
