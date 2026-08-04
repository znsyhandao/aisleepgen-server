# -*- coding: utf-8 -*-
"""Parse Mendeley XLSX to check demographics and PSQI"""
import openpyxl, os, json

xlsx_path = r'D:\AISleepGen_Optimized\data\edf\mendeley_insomnia\PSG_Psycho_Normal.xlsx'
wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
interesting = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ===')
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(list(row))
        if i < 5:
            print(f'  R{i}: {list(row)[:20]}')
    interesting[sheet_name] = rows

# Look for PSQI, insomnia diagnosis, or group labels
for sheet_name, rows in interesting.items():
    all_text = str(rows)
    if 'PSQI' in all_text or 'insomnia' in all_text.lower() or 'group' in all_text.lower():
        print(f'\n[FOUND KEYWORDS in {sheet_name}]')
        for r in rows[:20]:
            print(f'  {r[:15]}')

wb.close()
